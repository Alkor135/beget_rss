import sqlite3
import pandas as pd
import random
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

# --- Пути ---
db_path = Path(r"C:\Users\Alkor\gd\data_quote_db\RTS_days.sqlite")
output_xlsx = Path(__file__).parent / "rts_strategy_backtest_5.xlsx"

# --- Чтение данных из дневной БД ---
conn = sqlite3.connect(db_path)
df = pd.read_sql("SELECT TRADEDATE, OPEN, CLOSE FROM Futures ORDER BY TRADEDATE ASC", conn)
conn.close()

# --- Фактическое направление свечи ---
def candle_direction(row):
    if row["CLOSE"] > row["OPEN"]:
        return "UP"
    elif row["CLOSE"] < row["OPEN"]:
        return "DOWN"
    else:
        return "FLAT"

df["FACT_DIR"] = df.apply(candle_direction, axis=1)

# --- Списки для результатов ---
predictions = []
matches = []
results = []
cum_results = []

cum_profit = 0

# --- Основная логика стратегии ---
for i in range(len(df)):
    if i < 5:
        predictions.append(None)
        matches.append(None)
        results.append(None)
        cum_results.append(None)
    else:
        past_dirs = df.loc[i-5:i-1, "FACT_DIR"].tolist()
        pred = random.choice(past_dirs)
        predictions.append(pred)

        fact = df.loc[i, "FACT_DIR"]

        match = (pred == fact)
        matches.append(match)

        open_price = df.loc[i, "OPEN"]
        close_price = df.loc[i, "CLOSE"]

        if pred == "UP":
            res = close_price - open_price
        elif pred == "DOWN":
            res = open_price - close_price
        else:
            res = 0

        results.append(res)

        cum_profit += res
        cum_results.append(cum_profit)

# --- Результаты по сделкам ---
df["PREDICT"] = predictions
df["MATCH"] = matches
df["TRADE_RESULT"] = results
df["CUM_RESULT"] = cum_results

# --- Сводка ---
valid_trades = df["TRADE_RESULT"].dropna()
summary = {
    "Количество сделок": [len(valid_trades)],
    "Угадано (%)": [df["MATCH"].dropna().mean() * 100],
    "Общая прибыль": [valid_trades.sum()],
    "Средний результат на сделку": [valid_trades.mean()],
    "Максимальная просадка": [valid_trades.cumsum().min()],
    "Максимальный результат": [valid_trades.cumsum().max()]
}
summary_df = pd.DataFrame(summary)

# --- Сохранение в Excel ---
with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Backtest", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

# --- Добавление графика в Excel ---
wb = load_workbook(output_xlsx)
ws_backtest = wb["Backtest"]
ws_summary = wb["Summary"]

# Построим график по колонке CUM_RESULT
max_row = ws_backtest.max_row
chart = LineChart()
chart.title = "Кумулятивный результат"
chart.y_axis.title = "Прибыль"
chart.x_axis.title = "Номер сделки"

values = Reference(ws_backtest, min_col=ws_backtest.max_column, min_row=2, max_row=max_row)  # последний столбец = CUM_RESULT
chart.add_data(values, titles_from_data=False)

ws_summary.add_chart(chart, "A10")  # вставляем график ниже таблицы сводки

wb.save(output_xlsx)

print(f"Бектест завершён. Результаты и график сохранены в {output_xlsx}")
