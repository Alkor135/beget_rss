import sqlite3
import pandas as pd
import random
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series

# --- Пути ---
db_path = Path(r"C:\Users\Alkor\gd\data_quote_db\RTS_days.sqlite")
output_xlsx = Path(__file__).parent / "rts_strategy_backtest_multi.xlsx"

# --- Чтение данных из дневной БД ---
conn = sqlite3.connect(db_path)
df_base = pd.read_sql("SELECT TRADEDATE, OPEN, CLOSE FROM Futures ORDER BY TRADEDATE ASC", conn)
conn.close()

# --- Фактическое направление свечи ---
def candle_direction(row):
    if row["CLOSE"] > row["OPEN"]:
        return "UP"
    elif row["CLOSE"] < row["OPEN"]:
        return "DOWN"
    else:
        return "FLAT"

df_base["FACT_DIR"] = df_base.apply(candle_direction, axis=1)

# --- Диапазон параметров ---
n_past_range = range(2, 11)

# --- Хранилища результатов ---
summary_list = []
multi_results = pd.DataFrame({"TRADEDATE": df_base["TRADEDATE"]})

for N_PAST in n_past_range:
    predictions, matches, results, cum_results = [], [], [], []
    cum_profit = 0

    for i in range(len(df_base)):
        if i < N_PAST:
            predictions.append(None)
            matches.append(None)
            results.append(None)
            cum_results.append(None)
        else:
            past_dirs = df_base.loc[i-N_PAST:i-1, "FACT_DIR"].tolist()
            pred = random.choice(past_dirs)
            predictions.append(pred)

            fact = df_base.loc[i, "FACT_DIR"]
            match = (pred == fact)
            matches.append(match)

            open_price = df_base.loc[i, "OPEN"]
            close_price = df_base.loc[i, "CLOSE"]

            if pred == "UP":
                res = close_price - open_price
            elif pred == "DOWN":
                res = open_price - close_price
            else:
                res = 0

            results.append(res)
            cum_profit += res
            cum_results.append(cum_profit)

    # Добавляем колонку с кумулятивным результатом
    multi_results[f"CUM_{N_PAST}"] = cum_results

    # Сводка по прогону
    valid_trades = pd.Series(results).dropna()
    summary_list.append({
        "N_PAST": N_PAST,
        "Количество сделок": len(valid_trades),
        "Угадано (%)": pd.Series(matches).dropna().mean() * 100,
        "Общая прибыль": valid_trades.sum(),
        "Средний результат": valid_trades.mean(),
        "Максимальная просадка": valid_trades.cumsum().min(),
        "Максимальный результат": valid_trades.cumsum().max()
    })

summary_df = pd.DataFrame(summary_list)

# --- Сохранение в Excel ---
with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
    multi_results.to_excel(writer, sheet_name="BacktestMulti", index=False)
    summary_df.to_excel(writer, sheet_name="Summary", index=False)

# --- Добавление общего графика ---
wb = load_workbook(output_xlsx)
ws_multi = wb["BacktestMulti"]
ws_summary = wb["Summary"]

max_row = ws_multi.max_row
max_col = ws_multi.max_column

chart = LineChart()
chart.title = "Кумулятивный результат для разных N_PAST"
chart.y_axis.title = "Прибыль"
chart.x_axis.title = "Дата"

# Берём категории = даты (первая колонка TRADEDATE)
cats = Reference(ws_multi, min_col=1, min_row=2, max_row=max_row)

# Добавляем все серии кроме первой колонки
for col in range(2, max_col + 1):
    values = Reference(ws_multi, min_col=col, min_row=2, max_row=max_row)
    series = Series(values, title=ws_multi.cell(row=1, column=col).value)
    chart.series.append(series)

chart.set_categories(cats)  # <-- добавили даты на ось X

ws_summary.add_chart(chart, "A12")

wb.save(output_xlsx)

print(f"Мульти-бектест завершён. Результаты и график сохранены в {output_xlsx}")
